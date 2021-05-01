from django.shortcuts import render
from .models import Datos
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

class ReporteDatosExcel(TemplateView):
    def get(self,request,*args,**kwargs):
         #Obtenemos todas las personas de nuestra base de datos
        query6 = Datos.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['D1'].font=Font(name='Calibri', size=11, bold=True)
        ws['D1'] = 'REPORTE DE DATOS EN DJANGO 2.0'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('D1:G1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws.row_dimensions[3].height = 25
        ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['A3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['A3'].font=Font(name='Calibri', size=11, bold=True)
        ws['A3'] = 'Nombres'        


        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['B3'].font=Font(name='Calibri', size=11, bold=True)
        ws['B3'] = 'Apellidos'


        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['C3'].font=Font(name='Calibri', size=11, bold=True)
        ws['C3'] = 'Edad'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type='solid')
        ws['D3'].font=Font(name='Calibri', size=11, bold=True)
        ws['D3'] = 'Dirección'  

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        cont = 4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for que in query6:
            ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=1).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=1).font = Font(name='Calibri', size=8)
            ws.cell(row=cont,column=1).value = que.nombre
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=cont,column=2).value = que.apellidos
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=8)
            ws.cell(row=cont,column=3).value = que.edad
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=8) 
            ws.cell(row=cont,column=4).value = que.direccion            
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteDatosExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

